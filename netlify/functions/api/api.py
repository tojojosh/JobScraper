"""
Netlify Serverless Function – API handler for UK Skilled Jobs Portal.

This single function handles ALL /api/* routes.  Netlify redirects
/api/* to /.netlify/functions/api, and this handler parses the
original request path to dispatch to the correct logic.

Database: Connects via DATABASE_URL env var (PostgreSQL recommended).
Falls back to SQLite in /tmp/ for testing (data won't persist across
cold starts).
"""
import json
import os
import csv
import io
import base64
from datetime import date, datetime, timedelta

from sqlalchemy import (
    create_engine, Column, Integer, String, Date, DateTime,
    Boolean, Text, UniqueConstraint, or_, func,
)
from sqlalchemy.orm import declarative_base, sessionmaker


# ═══════════════════════════════════════════════════════════════════
# DATABASE SETUP
# ═══════════════════════════════════════════════════════════════════

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:////tmp/jobs.db')

# Neon / Supabase / Heroku use postgres:// but SQLAlchemy needs postgresql://
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=300)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()


# ── ORM Models ───────────────────────────────────────────────────

class Job(Base):
    __tablename__ = 'jobs'

    id = Column(Integer, primary_key=True)
    title = Column(String(500), nullable=False)
    company = Column(String(300), nullable=False)
    location = Column(String(300), nullable=False)
    category = Column(String(200), nullable=True)
    experience_level = Column(String(100), nullable=True)
    job_type = Column(String(100), nullable=True)
    salary = Column(String(200), nullable=True)
    url = Column(String(2000), nullable=False)
    url_hash = Column(String(64), nullable=False, index=True)
    source = Column(String(100), nullable=False)
    scrape_date = Column(Date, nullable=False, index=True)
    first_seen_date = Column(Date, nullable=False)
    last_seen_date = Column(Date, nullable=False)
    created_at = Column(DateTime)

    __table_args__ = (
        UniqueConstraint('url_hash', 'scrape_date', name='uq_job_url_date'),
    )


class TargetCompany(Base):
    __tablename__ = 'target_companies'

    id = Column(Integer, primary_key=True)
    name = Column(String(300), nullable=False, unique=True)
    career_url = Column(String(2000), nullable=True)
    active = Column(Boolean, default=True)


class ScrapeRun(Base):
    __tablename__ = 'scrape_runs'

    id = Column(Integer, primary_key=True)
    run_date = Column(Date, nullable=False)
    status = Column(String(20), nullable=False, default='running')
    jobs_found = Column(Integer, default=0)
    new_jobs = Column(Integer, default=0)
    duplicates = Column(Integer, default=0)
    failed_sources = Column(Integer, default=0)
    started_at = Column(DateTime)
    completed_at = Column(DateTime, nullable=True)
    log = Column(Text, nullable=True)


# ── DB helpers ───────────────────────────────────────────────────

_db_initialized = False


def _init_db():
    global _db_initialized
    if not _db_initialized:
        Base.metadata.create_all(bind=engine)
        _db_initialized = True


def _get_session():
    _init_db()
    return SessionLocal()


# ═══════════════════════════════════════════════════════════════════
# HANDLER  (Netlify entry point)
# ═══════════════════════════════════════════════════════════════════

def handler(event, context):
    """Main entry point – dispatches to route handlers."""

    # Handle CORS preflight
    if event.get('httpMethod') == 'OPTIONS':
        return _cors_response()

    path = event.get('path', '')
    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}

    # Normalise route: strip function path and /api/ prefix
    route = path
    for prefix in ['/.netlify/functions/api', '/api']:
        if route.startswith(prefix):
            route = route[len(prefix):]
            break
    route = route.strip('/')

    try:
        # ── Route dispatch ──────────────────────────────────────
        if route == 'jobs' and method == 'GET':
            return _get_jobs(params)

        elif route == 'jobs/export/csv':
            return _export_csv(params)

        elif route == 'jobs/export/excel':
            return _export_excel(params)

        elif route == 'jobs/export/json':
            return _export_json_range(params)

        elif route.startswith('jobs/daily-json/'):
            date_str = route.split('/')[-1]
            return _daily_json(date_str)

        elif route == 'stats':
            return _get_stats(params)

        elif route == 'dates':
            return _get_dates()

        elif route == 'companies':
            return _get_companies()

        elif route == 'scrape' and method == 'POST':
            return _scrape_unavailable()

        elif route == 'scrape/status':
            return _get_scrape_status()

        else:
            return _json_response(404, {'error': f'Not found: /api/{route}'})

    except Exception as exc:
        print(f"[ERROR] {method} {path}: {exc}")
        return _json_response(500, {'error': str(exc)})


# ═══════════════════════════════════════════════════════════════════
# RESPONSE HELPERS
# ═══════════════════════════════════════════════════════════════════

_CORS_HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Headers': 'Content-Type',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
}


def _cors_response():
    return {'statusCode': 204, 'headers': _CORS_HEADERS, 'body': ''}


def _json_response(status, body):
    return {
        'statusCode': status,
        'headers': {**_CORS_HEADERS, 'Content-Type': 'application/json'},
        'body': json.dumps(body, default=str),
    }


def _file_response(body_bytes, content_type, filename):
    return {
        'statusCode': 200,
        'headers': {
            **_CORS_HEADERS,
            'Content-Type': content_type,
            'Content-Disposition': f'attachment; filename={filename}',
        },
        'body': base64.b64encode(body_bytes).decode('utf-8'),
        'isBase64Encoded': True,
    }


# ═══════════════════════════════════════════════════════════════════
# DATE HELPERS
# ═══════════════════════════════════════════════════════════════════

def _parse_date(val, default=None):
    if not val:
        return default
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except ValueError:
        return default


# ═══════════════════════════════════════════════════════════════════
# SHARED QUERY BUILDER
# ═══════════════════════════════════════════════════════════════════

def _build_filtered_query(session, params):
    """Build a filtered + sorted Job query from request params."""
    query = session.query(Job)

    date_from = _parse_date(params.get('date_from'), date.today() - timedelta(days=7))
    date_to = _parse_date(params.get('date_to'), date.today())
    query = query.filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)

    search = (params.get('search') or '').strip()
    if search:
        term = f'%{search}%'
        query = query.filter(or_(
            Job.title.ilike(term),
            Job.company.ilike(term),
            Job.location.ilike(term),
        ))

    source = (params.get('source') or '').strip()
    if source:
        query = query.filter(Job.source == source)

    sort_by = params.get('sort_by', 'scrape_date')
    sort_order = params.get('sort_order', 'desc')
    sort_map = {
        'scrape_date': Job.scrape_date,
        'company': Job.company,
        'title': Job.title,
        'location': Job.location,
        'source': Job.source,
    }
    col = sort_map.get(sort_by, Job.scrape_date)
    query = query.order_by(col.asc() if sort_order == 'asc' else col.desc())

    return query


# ═══════════════════════════════════════════════════════════════════
# MODEL SERIALIZERS
# ═══════════════════════════════════════════════════════════════════

def _job_dict(j):
    return {
        'id': j.id,
        'title': j.title,
        'company': j.company,
        'location': j.location,
        'category': j.category,
        'experience_level': j.experience_level,
        'job_type': j.job_type,
        'salary': j.salary,
        'url': j.url,
        'source': j.source,
        'scrape_date': j.scrape_date.isoformat() if j.scrape_date else None,
        'first_seen_date': j.first_seen_date.isoformat() if j.first_seen_date else None,
        'last_seen_date': j.last_seen_date.isoformat() if j.last_seen_date else None,
    }


def _job_export(j):
    return {
        'title': j.title,
        'company': j.company,
        'location': j.location,
        'category': j.category,
        'experience_level': j.experience_level,
        'job_type': j.job_type,
        'salary': j.salary,
        'url': j.url,
    }


# ═══════════════════════════════════════════════════════════════════
# ROUTE HANDLERS
# ═══════════════════════════════════════════════════════════════════

# ── GET /api/jobs ────────────────────────────────────────────────

def _get_jobs(params):
    session = _get_session()
    try:
        query = _build_filtered_query(session, params)

        page = max(1, int(params.get('page', 1)))
        page_size = min(int(params.get('page_size', 50)), 200)
        total = query.count()
        total_pages = max(1, (total + page_size - 1) // page_size)

        jobs = query.offset((page - 1) * page_size).limit(page_size).all()

        return _json_response(200, {
            'jobs': [_job_dict(j) for j in jobs],
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_prev': page > 1,
        })
    finally:
        session.close()


# ── GET /api/jobs/export/csv ─────────────────────────────────────

def _export_csv(params):
    session = _get_session()
    try:
        jobs = _build_filtered_query(session, params).all()

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            'S/NO', 'Company Name', 'Job Title', 'Job Link', 'Location',
            'Salary', 'Category', 'Experience Level', 'Job Type',
            'Source', 'Scrape Date',
        ])
        for idx, j in enumerate(jobs, 1):
            writer.writerow([
                idx, j.company, j.title, j.url, j.location,
                j.salary or '', j.category or '', j.experience_level or '',
                j.job_type or '', j.source,
                j.scrape_date.isoformat() if j.scrape_date else '',
            ])

        return _file_response(
            buf.getvalue().encode('utf-8'),
            'text/csv',
            f'uk_jobs_{date.today().isoformat()}.csv',
        )
    finally:
        session.close()


# ── GET /api/jobs/export/excel ───────────────────────────────────

def _export_excel(params):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = _get_session()
    try:
        jobs = _build_filtered_query(session, params).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'UK Jobs'

        headers = [
            'S/NO', 'Company Name', 'Job Title', 'Job Link', 'Location',
            'Salary', 'Category', 'Experience Level', 'Job Type',
            'Source', 'Scrape Date',
        ]
        ws.append(headers)

        hfont = Font(bold=True, color='FFFFFF')
        hfill = PatternFill(start_color='1a73e8', end_color='1a73e8',
                            fill_type='solid')
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')

        for idx, j in enumerate(jobs, 1):
            ws.append([
                idx, j.company, j.title, j.url, j.location,
                j.salary or '', j.category or '', j.experience_level or '',
                j.job_type or '', j.source,
                j.scrape_date.isoformat() if j.scrape_date else '',
            ])

        for col_cells in ws.columns:
            max_len = 0
            letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except TypeError:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)

        return _file_response(
            buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            f'uk_jobs_{date.today().isoformat()}.xlsx',
        )
    finally:
        session.close()


# ── GET /api/jobs/export/json (date range) ───────────────────────

def _export_json_range(params):
    date_from = _parse_date(params.get('date_from'))
    date_to = _parse_date(params.get('date_to'))

    if not date_from or not date_to:
        return _json_response(400, {
            'error': 'Both date_from and date_to required (YYYY-MM-DD).',
        })
    if date_from > date_to:
        return _json_response(400, {
            'error': 'date_from must be on or before date_to.',
        })

    session = _get_session()
    try:
        jobs = (
            session.query(Job)
            .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
            .order_by(Job.scrape_date.desc(), Job.company)
            .all()
        )
        data = json.dumps([_job_export(j) for j in jobs], indent=2)
        fname = f'jobs_{date_from.isoformat()}_to_{date_to.isoformat()}.json'

        return {
            'statusCode': 200,
            'headers': {
                **_CORS_HEADERS,
                'Content-Type': 'application/json',
                'Content-Disposition': f'attachment; filename={fname}',
            },
            'body': data,
        }
    finally:
        session.close()


# ── GET /api/jobs/daily-json/<date> ──────────────────────────────

def _daily_json(date_str):
    target = _parse_date(date_str)
    if not target:
        return _json_response(400, {
            'error': 'Invalid date format. Use YYYY-MM-DD',
        })

    session = _get_session()
    try:
        jobs = session.query(Job).filter_by(scrape_date=target).all()
        data = json.dumps([_job_export(j) for j in jobs], indent=2)

        return {
            'statusCode': 200,
            'headers': {
                **_CORS_HEADERS,
                'Content-Type': 'application/json',
                'Content-Disposition': f'attachment; filename=jobs_{date_str}.json',
            },
            'body': data,
        }
    finally:
        session.close()


# ── GET /api/stats ───────────────────────────────────────────────

def _get_stats(params):
    date_from = _parse_date(
        params.get('date_from'), date.today() - timedelta(days=7),
    )
    date_to = _parse_date(params.get('date_to'), date.today())

    session = _get_session()
    try:
        total_jobs = (
            session.query(func.count(Job.id))
            .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
            .scalar()
        )

        unique_companies = (
            session.query(func.count(func.distinct(Job.company)))
            .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
            .scalar()
        )

        sources = (
            session.query(Job.source, func.count(Job.id))
            .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
            .group_by(Job.source)
            .all()
        )

        last_run = (
            session.query(ScrapeRun)
            .order_by(ScrapeRun.started_at.desc())
            .first()
        )

        return _json_response(200, {
            'total_jobs': total_jobs or 0,
            'unique_companies': unique_companies or 0,
            'sources': {s: c for s, c in sources},
            'date_range': {
                'from': date_from.isoformat(),
                'to': date_to.isoformat(),
            },
            'last_run': {
                'date': last_run.run_date.isoformat(),
                'status': last_run.status,
                'jobs_found': last_run.jobs_found,
                'new_jobs': last_run.new_jobs,
            } if last_run else None,
        })
    finally:
        session.close()


# ── GET /api/dates ───────────────────────────────────────────────

def _get_dates():
    session = _get_session()
    try:
        rows = (
            session.query(Job.scrape_date, func.count(Job.id))
            .group_by(Job.scrape_date)
            .order_by(Job.scrape_date.desc())
            .all()
        )
        return _json_response(200, {
            'dates': [
                {'date': d.isoformat(), 'count': c} for d, c in rows
            ],
        })
    finally:
        session.close()


# ── GET /api/companies ───────────────────────────────────────────

def _get_companies():
    session = _get_session()
    try:
        rows = (
            session.query(TargetCompany)
            .filter_by(active=True)
            .order_by(TargetCompany.name)
            .all()
        )
        return _json_response(200, {
            'companies': [{'id': c.id, 'name': c.name} for c in rows],
        })
    finally:
        session.close()


# ── POST /api/scrape ─────────────────────────────────────────────

def _scrape_unavailable():
    return _json_response(200, {
        'status': 'unavailable',
        'error': (
            'Scraping is not available in the Netlify deployment due to '
            'serverless timeout limits. Run the scraper locally with '
            '"python app.py" connected to the same DATABASE_URL.'
        ),
    })


# ── GET /api/scrape/status ───────────────────────────────────────

def _get_scrape_status():
    session = _get_session()
    try:
        last_run = (
            session.query(ScrapeRun)
            .order_by(ScrapeRun.started_at.desc())
            .first()
        )
        if not last_run:
            return _json_response(200, {'status': 'no_runs'})

        return _json_response(200, {
            'id': last_run.id,
            'date': last_run.run_date.isoformat(),
            'status': last_run.status,
            'jobs_found': last_run.jobs_found,
            'new_jobs': last_run.new_jobs,
            'duplicates': last_run.duplicates,
            'failed_sources': last_run.failed_sources,
            'started_at': (last_run.started_at.isoformat()
                           if last_run.started_at else None),
            'completed_at': (last_run.completed_at.isoformat()
                             if last_run.completed_at else None),
        })
    finally:
        session.close()
