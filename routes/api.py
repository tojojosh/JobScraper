"""REST API endpoints for the jobs portal."""
import csv
import io
import json
from datetime import date, datetime, timedelta

from flask import Blueprint, request, jsonify, Response, send_file
from models import db, Job, ScrapeRun, TargetCompany

api_bp = Blueprint('api', __name__, url_prefix='/api')


# ── Helpers ──────────────────────────────────────────────────────────

def _parse_date(date_str, default=None):
    if not date_str:
        return default
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return default


def _build_jobs_query():
    """Return a SQLAlchemy query filtered/sorted by request args."""
    query = Job.query

    # Date range
    date_from = _parse_date(
        request.args.get('date_from'),
        date.today() - timedelta(days=7),
    )
    date_to = _parse_date(request.args.get('date_to'), date.today())
    query = query.filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)

    # Keyword search (title / company / location)
    search = request.args.get('search', '').strip()
    if search:
        term = f'%{search}%'
        query = query.filter(
            db.or_(
                Job.title.ilike(term),
                Job.company.ilike(term),
                Job.location.ilike(term),
            )
        )

    # Company filter
    company = request.args.get('company', '').strip()
    if company:
        query = query.filter(Job.company.ilike(f'%{company}%'))

    # Source filter
    source = request.args.get('source', '').strip()
    if source:
        query = query.filter(Job.source == source)

    # Sorting
    sort_by = request.args.get('sort_by', 'scrape_date')
    sort_order = request.args.get('sort_order', 'desc')
    sort_columns = {
        'scrape_date': Job.scrape_date,
        'company': Job.company,
        'title': Job.title,
        'location': Job.location,
        'source': Job.source,
    }
    sort_col = sort_columns.get(sort_by, Job.scrape_date)
    query = query.order_by(sort_col.asc() if sort_order == 'asc' else sort_col.desc())

    return query


# ── Jobs list (paginated) ───────────────────────────────────────────

@api_bp.route('/jobs')
def get_jobs():
    query = _build_jobs_query()

    page = max(1, request.args.get('page', 1, type=int))
    page_size = min(request.args.get('page_size', 50, type=int), 200)

    pagination = query.paginate(page=page, per_page=page_size, error_out=False)

    return jsonify({
        'jobs': [job.to_dict() for job in pagination.items],
        'total': pagination.total,
        'page': pagination.page,
        'page_size': page_size,
        'total_pages': pagination.pages,
        'has_next': pagination.has_next,
        'has_prev': pagination.has_prev,
    })


# ── CSV Export ───────────────────────────────────────────────────────

@api_bp.route('/jobs/export/csv')
def export_csv():
    jobs = _build_jobs_query().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        'S/NO', 'Company Name', 'Job Title', 'Job Link', 'Location',
        'Salary', 'Category', 'Experience Level', 'Job Type', 'Source', 'Scrape Date',
    ])
    for idx, job in enumerate(jobs, 1):
        writer.writerow([
            idx, job.company, job.title, job.url, job.location,
            job.salary or '', job.category or '', job.experience_level or '',
            job.job_type or '', job.source, job.scrape_date.isoformat(),
        ])

    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition':
                f'attachment; filename=uk_jobs_{date.today().isoformat()}.csv'
        },
    )


# ── Excel Export ─────────────────────────────────────────────────────

@api_bp.route('/jobs/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    jobs = _build_jobs_query().all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'UK Jobs'

    headers = [
        'S/NO', 'Company Name', 'Job Title', 'Job Link', 'Location',
        'Salary', 'Category', 'Experience Level', 'Job Type', 'Source', 'Scrape Date',
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for idx, job in enumerate(jobs, 1):
        ws.append([
            idx, job.company, job.title, job.url, job.location,
            job.salary or '', job.category or '', job.experience_level or '',
            job.job_type or '', job.source, job.scrape_date.isoformat(),
        ])

    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except TypeError:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'uk_jobs_{date.today().isoformat()}.xlsx',
    )


# ── Daily JSON ───────────────────────────────────────────────────────

@api_bp.route('/jobs/daily-json/<date_str>')
def daily_json(date_str):
    target = _parse_date(date_str)
    if not target:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    jobs = Job.query.filter_by(scrape_date=target).all()
    data = [job.to_json_export() for job in jobs]

    return Response(
        json.dumps(data, indent=2),
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename=jobs_{date_str}.json'
        },
    )


# ── Statistics ───────────────────────────────────────────────────────

@api_bp.route('/stats')
def get_stats():
    date_from = _parse_date(
        request.args.get('date_from'),
        date.today() - timedelta(days=7),
    )
    date_to = _parse_date(request.args.get('date_to'), date.today())

    total_jobs = Job.query.filter(
        Job.scrape_date >= date_from,
        Job.scrape_date <= date_to,
    ).count()

    unique_companies = (
        db.session.query(db.func.count(db.func.distinct(Job.company)))
        .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
        .scalar()
    )

    sources = (
        db.session.query(Job.source, db.func.count(Job.id))
        .filter(Job.scrape_date >= date_from, Job.scrape_date <= date_to)
        .group_by(Job.source)
        .all()
    )

    last_run = ScrapeRun.query.order_by(ScrapeRun.started_at.desc()).first()

    return jsonify({
        'total_jobs': total_jobs,
        'unique_companies': unique_companies or 0,
        'sources': {s: c for s, c in sources},
        'date_range': {
            'from': date_from.isoformat(),
            'to': date_to.isoformat(),
        },
        'last_run': {
            'date': last_run.run_date.isoformat(),
            'status': last_run.status,
            'jobs_found': last_run.jobs_found,
            'new_jobs': last_run.new_jobs,
        } if last_run else None,
    })


# ── Available scrape dates ───────────────────────────────────────────

@api_bp.route('/dates')
def get_dates():
    rows = (
        db.session.query(Job.scrape_date, db.func.count(Job.id))
        .group_by(Job.scrape_date)
        .order_by(Job.scrape_date.desc())
        .all()
    )
    return jsonify({
        'dates': [{'date': d.isoformat(), 'count': c} for d, c in rows],
    })


# ── Manual scrape trigger ────────────────────────────────────────────

@api_bp.route('/scrape', methods=['POST'])
def trigger_scrape():
    from flask import current_app
    from scraper.engine import ScrapingEngine

    engine = ScrapingEngine(current_app._get_current_object())
    result = engine.run()
    return jsonify(result)


# ── Scrape status ────────────────────────────────────────────────────

@api_bp.route('/scrape/status')
def scrape_status():
    last_run = ScrapeRun.query.order_by(ScrapeRun.started_at.desc()).first()
    if not last_run:
        return jsonify({'status': 'no_runs'})

    return jsonify({
        'id': last_run.id,
        'date': last_run.run_date.isoformat(),
        'status': last_run.status,
        'jobs_found': last_run.jobs_found,
        'new_jobs': last_run.new_jobs,
        'duplicates': last_run.duplicates,
        'failed_sources': last_run.failed_sources,
        'started_at': last_run.started_at.isoformat() if last_run.started_at else None,
        'completed_at': last_run.completed_at.isoformat() if last_run.completed_at else None,
    })


# ── Target companies management ──────────────────────────────────────

@api_bp.route('/companies')
def get_companies():
    companies = TargetCompany.query.filter_by(active=True).order_by(TargetCompany.name).all()
    return jsonify({
        'companies': [{'id': c.id, 'name': c.name} for c in companies],
    })
